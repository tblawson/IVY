# -*- coding: utf-8 -*-
"""
Created on Wed Jun 24 09:36:42 2015

DEVELOPMENT VERSION

acquisition.py:
Thread class that executes processing.
Contains definitions for usual __init__() and run() methods
 AND an abort() method. The Run() method forms the core of the
 procedure - any changes to the way the measurements are taken
 should be made here, and within included subroutines.

@author: t.lawson
"""

import wx
from threading import Thread
import datetime as dt
import time
import numpy as np
import logging
import json

from openpyxl.styles import Font, Border, Side

import IVY_events as evts
import devices

NREADS = 20
TEST_V_OUT = [0.1, 1, 10]  # O/P test voltage selection
NODES = ['V1', 'V2']  # Input node selection
TEST_V_MASK = [0, -1, 1, 0]  # Polarity / zero selection
I_MIN = 1e-11  # 10 pA (S/N issues, noise limit)
I_MAX = 0.01  # 10 mA (Opamp overheating limit)
V1_MIN = 0.01  # 10 mV (S/N issues, noise limit)
V1_MAX = 10  # 10 V (Opamp output limit)
P_MAX = 480  # Maximum progress (20 measurement-cycles * 24 rows)

logger = logging.getLogger(__name__)


class AqnThread(Thread):
    """Acquisition Thread Class."""
    def __init__(self, parent):
        # This runs when an instance of the class is created
        Thread.__init__(self)
        self.RunPage = parent
        self.SetupPage = self.RunPage.GetParent().GetPage(0)
        self.PlotPage = self.RunPage.GetParent().GetPage(2)
        self.CalcPage = self.RunPage.GetParent().GetPage(3)
        self.TopLevel = self.RunPage.GetTopLevelParent()
        self.Comment = self.RunPage.Comment.GetValue()

        self._want_abort = 0

        self.V12Data = {'V1': [], 'V2': []}
        self.V3Data = []
        self.Times = []
        self.V12m = {'V1': 0, 'V2': 0}
        self.V12sd = {'V1': 0, 'V2': 0}

        # Dictionary for accumulating data for this run:
        self.run_dict = {'Comment': self.Comment,
                         'Rs': self.RunPage.Rs_val,
                         'DUC_G': float(self.RunPage.DUCgain.GetValue()),
                         'Instruments': {},
                         'Nreads': NREADS,
                         'Date_time': [],
                         'Node': [],
                         'Nom_Vout': [],
                         'OP_V': {'val': [], 'sd': []},
                         'OPrange': [],
                         'IP_V': {'val': [], 'sd': []},
                         'IPrange': [],
                         'T_GMH': [],
                         'Pt_DVM': [],
                         'Room_conds': {'T': [], 'P': [], 'RH': []},
                         }

        print'Role -> Instrument:'
        print'------------------------------'
        logger.info('Role -> Instrument:')
        # Print all device objects
        for r in devices.ROLES_INSTR.keys():
            if r == 'IVbox':
                d = 'IV_box'
            else:
                d = devices.ROLES_INSTR[r].Descr
            sub_dict = {r: d}
            self.run_dict['Instruments'].update(sub_dict)
            print'%s \t-> %s' % (devices.INSTR_DATA[d]['role'], d)
            logger.info('%s \t-> %s', devices.INSTR_DATA[d]['role'], d)
#            if r != devices.INSTR_DATA[d]['role']:
#                devices.INSTR_DATA[d]['role'] = r
#                print'Role data corrected to:', r, '->', d
#                logger.info('Role data corrected to: %s -> %s', r, d)

#        # Get filename of Excel file
#        self.xlfilename = self.SetupPage.XLFile.GetValue()  # Full path
#        self.path_components = self.xlfilename.split('\\')
#        self.directory = '\\'.join(self.path_components[0:-1])

#        # Find existing workbook
#        self.wb_io = self.SetupPage.wb
#        self.ws = self.wb_io.get_sheet_by_name('Data')
#
#        # read start row number from Excel file
#        self.start_row = self.ws['B1'].value

#        strt_ev = evts.StartRowEvent(row=self.start_row)
#        wx.PostEvent(self.RunPage, strt_ev)

        self.settle_time = self.RunPage.SettleDel.GetValue()

        # Local record of GMH ports and addresses
        self.GMH1Demo_status = devices.ROLES_INSTR['GMH'].demo
        self.GMH1Port = devices.ROLES_INSTR['GMH'].addr

        self.start()  # Starts the thread running on creation

    def run(self):
        '''
        Run Worker Thread.
        This is where all the important stuff goes, in a repeated cycle.
        '''
        print'\nRUN START...\n'
        logger.info('RUN START...')

        # Set button availability
        self.RunPage.StopBtn.Enable(True)
        self.RunPage.StartBtn.Enable(False)

        # Clear plots
        clr_plot_ev = evts.ClearPlotEvent()
        wx.PostEvent(self.PlotPage, clr_plot_ev)

#        self.WriteHeadings()

        stat_ev = evts.StatusEvent(msg='AqnThread.run():', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        stat_ev = evts.StatusEvent(msg='Waiting to settle...', field=1)
        wx.PostEvent(self.TopLevel, stat_ev)

        time.sleep(self.settle_time)

        # Initialise all instruments (doesn't open GMH sensors yet)
        if self._want_abort:
            self.AbortRun()
            return
        self.initialise()

        stat_ev = evts.StatusEvent(msg='',
                                   field='b')  # write to both status fields
        wx.PostEvent(self.TopLevel, stat_ev)

        if self._want_abort:
            self.AbortRun()
            return
        stat_ev = evts.StatusEvent(msg=' Post-initialise delay (3s)', field=1)
        wx.PostEvent(self.TopLevel, stat_ev)
#        time.sleep(3)

#        self.WriteInstrAssignments()

#        row = self.start_row  # Start of data
        row = 1
        pbar = 0

        '''
        The following sequence progresses through three blocks with nominal
        output voltages of 0.1, 1 or 10 V. For each output voltage block
        the input DVM switches between the two nodes V1 and V2.
        For each input node, a mask is applied to the output voltage (and
        thus input V) causing the value to be set to 0 V, each polarity,
        then 0 V again.

        Nominal input voltage and current are calculated based on nominal
        DUC gain and Rs. If the nominal input current would be outside the
        scope {1e-11 A < I < 1e-3 A}, that 8-row block is skipped. If the
        nominal calculated input voltage, required to result in the prescribed
        output, would be outside the scope {0.01 V < V < 10 V}, that 8-row
        block is skipped.
        '''
        self.Rs = self.RunPage.Rs_val
        self.DUC_G = float(self.RunPage.DUCgain.GetValue())

        for abs_V3 in TEST_V_OUT:  # Loop over desired output voltages
            print'\nV3:', abs_V3, 'V'
            logger.info('V3: %d V', abs_V3)
            self.V1_nom = self.Rs*abs_V3/self.DUC_G
            self.I_nom = self.V1_nom/self.Rs
            if (abs(self.I_nom) <= I_MIN or abs(self.I_nom) >= I_MAX):
                I_nom_str = '(%.1g A)' % self.I_nom
                warning = '\nNominal I/P test-I outside scope! '+I_nom_str
                print warning
                logger.warning(warning)
                stat_ev = evts.StatusEvent(msg=warning, field=1)
                wx.PostEvent(self.TopLevel, stat_ev)
                pbar += 160
                Update = {'node': '-', 'Vm': 0, 'Vsc': 0, 'time': '-',
                          'row': row, 'progress': 100.0*pbar/P_MAX,
                          'end_flag': 0}

                update_ev = evts.DataEvent(ud=Update)
                wx.PostEvent(self.RunPage, update_ev)
                continue

            if abs(self.V1_nom) < V1_MIN or abs(self.V1_nom) > V1_MAX:
                V_nom_str = '(%.1g V)' % self.V1_nom
                warning = '\nNom. I/P test-V outside scope! '+V_nom_str
                print warning
                logger.warning(warning)
                stat_ev = evts.StatusEvent(msg=warning, field=1)
                wx.PostEvent(self.TopLevel, stat_ev)
                pbar += 160
                Update = {'node': '-', 'Vm': 0, 'Vsc': 0, 'time': '-',
                          'row': row, 'progress': 100.0*pbar/P_MAX,
                          'end_flag': 0}
                update_ev = evts.DataEvent(ud=Update)
                wx.PostEvent(self.RunPage, update_ev)
                continue

            Update = {'node': '-', 'Vm': 0, 'Vsc': 0, 'time': '-', 'row': row,
                      'progress': 100.0*pbar/P_MAX, 'end_flag': 0}
            update_ev = evts.DataEvent(ud=Update)
            wx.PostEvent(self.RunPage, update_ev)
            for node in NODES:  # Select input node (V1 then V2)
                self.SetNode(node)

                for V3_mask in TEST_V_MASK:
                    '''
                    Loop over {0,+,-,0} test voltages (assumes negative gain)
                    '''
                    self.Vout = abs_V3*V3_mask  # Nominal output
                    self.V1_set = -1.0*self.Vout*self.Rs/self.DUC_G
                    if abs(self.V1_set) == 0:
                        self.V1_set = 0.0
                    print'I/P test-V = %f \tO/P test-V = %f' % (self.V1_set,
                                                                self.Vout)
                    logger.info('I/P test-V = %f \tO/P test-V = %f',
                                self.V1_set, self.Vout)

                    stat_ev = evts.StatusEvent(msg='AqnThread.run():', field=0)
                    wx.PostEvent(self.TopLevel, stat_ev)
                    stat_ev = evts.StatusEvent(msg='I/P test-V = ' +
                                               str(self.V1_set) +
                                               '. O/P test-V = ' +
                                               str(self.Vout), field=1)
                    wx.PostEvent(self.TopLevel, stat_ev)

                    self.SetUpMeasThisRow(node)  # Clear F5520A errors & data
                    if self._want_abort:
                        self.AbortRun()
                        return
                    if not devices.ROLES_INSTR['SRC'].demo:
                        time.sleep(3)  # Wait 3s after checking F5520A error

                    '''
                    Set DVM ranges to suit voltages that they're
                    about to be exposed to:
                    '''
                    devices.ROLES_INSTR['DVM12'].SendCmd('DCV AUTO')
                    devices.ROLES_INSTR['DVM3'].SendCmd('DCV AUTO')
                    if not (devices.ROLES_INSTR['DVM12'].demo and
                            devices.ROLES_INSTR['DVM3'].demo):
                                time.sleep(0.5)  # Settle after setting range

                    print 'Aqn_thread.run(): masked V1_set =', self.V1_set
                    logger.info('masked V1_set = %f', self.V1_set)
                    self.RunPage.V1Setting.SetValue(str(self.V1_set))
                    if not devices.ROLES_INSTR['SRC'].demo:
                        time.sleep(0.5)  # Settle after setting V
                    if self.V1_set == 0:
                        devices.ROLES_INSTR['SRC'].Oper()  # Over-ride 0V STBY
                    if self._want_abort:
                        self.AbortRun()
                        return
                    if not devices.ROLES_INSTR['SRC'].demo:
                        time.sleep(30)  # wait after applying V and OPER mode.

#                    if self._want_abort:
#                        self.AbortRun()
#                        return
#                    time.sleep(5)  # wait 5s after setting voltage

                    # Prepare DVMs...
                    stat_ev = evts.StatusEvent(msg='Preparing DVMs...',
                                               field=1)
                    wx.PostEvent(self.TopLevel, stat_ev)

                    devices.ROLES_INSTR['DVM12'].SendCmd('LFREQ LINE')
                    devices.ROLES_INSTR['DVM3'].SendCmd('LFREQ LINE')
                    if self._want_abort:
                        self.AbortRun()
                        return
                    if not (devices.ROLES_INSTR['DVM12'].demo and
                            devices.ROLES_INSTR['DVM3'].demo):
                                time.sleep(3)

                    devices.ROLES_INSTR['DVM12'].SendCmd('AZERO ONCE')
                    devices.ROLES_INSTR['DVM3'].SendCmd('AZERO ON')
                    if self._want_abort:
                        self.AbortRun()
                        return
                    if not (devices.ROLES_INSTR['DVM12'].demo and devices.ROLES_INSTR['DVM3'].demo):
                        time.sleep(30)  # 30

                    status_msg = 'Making {0:d} measurements each of {1:s} and\
 V3 (V1_nom = {2:.2f} V)'.format(NREADS, node, self.V1_nom)
                    print status_msg
                    logger.info(status_msg)
                    stat_ev = evts.StatusEvent(msg=status_msg, field=1)
                    wx.PostEvent(self.TopLevel, stat_ev)

                    for n in range(NREADS):  # Acquire all V and t readings
                        self.MeasureV(node)
                        self.MeasureV('V3')
                        pbar += 1
                        Update = {'node': '-', 'Vm': 0, 'Vsd': 0, 'time': '-',
                                  'row': row, 'progress': 100.0*pbar/(P_MAX),
                                  'end_flag': 0}

                        update_ev = evts.DataEvent(ud=Update)
                        wx.PostEvent(self.RunPage, update_ev)
                        if self._want_abort:
                            self.AbortRun()
                            return
                    print'\n'
                    time.sleep(1)

                    assert len(self.V12Data[node]) == NREADS, 'Number of {0:s} readings != {1:d}!'.format(node, NREADS)
                    assert len(self.Times) == NREADS, 'Number of timestamps != {1:d}!'.format(NREADS)
                    self.tm = dt.datetime.fromtimestamp(np.mean(self.Times)).strftime("%d/%m/%Y %H:%M:%S")
                    self.V12m[node] = np.mean(self.V12Data[node])
                    self.V12sd[node] = np.std(self.V12Data[node], ddof=1)
                    meas_status = 'V12m[{0:s}] = {1:.6f}'.format(node, self.V12m[node])
                    print meas_status
                    logger.info(meas_status)
                    self.SetNode(node)
                    Update = {'node': node, 'Vm': self.V12m[node],
                              'Vsd': self.V12sd[node], 'time': self.tm,
                              'row': row, 'progress': 100.0*pbar/(P_MAX),
                              'end_flag': 0}

                    update_ev = evts.DataEvent(ud=Update)
                    wx.PostEvent(self.RunPage, update_ev)
                    IPrange = devices.ROLES_INSTR['DVM12'].SendCmd('RANGE?')
                    if not isinstance(IPrange, float):
                        IPrange = 0.0
                    self.IPrange = IPrange

                    time.sleep(2)  # Give user time to read vals before update

                    assert len(self.V3Data) == NREADS, 'Number of V3 readings != {1:d}!'.format(NREADS)
                    self.V3m = np.mean(self.V3Data)
                    self.V3sd = np.std(self.V3Data, ddof=1)
                    self.T = devices.ROLES_INSTR['GMH'].Measure('T')
                    OPrange = devices.ROLES_INSTR['DVM3'].SendCmd('RANGE?')
                    if not isinstance(OPrange, float):
                        OPrange = 0.0
                    self.OPrange = OPrange

                    self.SetNode('V3')
                    Update = {'node': 'V3', 'Vm': self.V3m, 'Vsd': self.V3sd,
                              'time': self.tm, 'row': row, 'end_flag': 0}
                    update_ev = evts.DataEvent(ud=Update)
                    wx.PostEvent(self.RunPage, update_ev)

                    if self._want_abort:
                        self.AbortRun()
                        return
                    stat_ev = evts.StatusEvent(msg="Post-acqisn. delay (5s)",
                                               field=1)
                    wx.PostEvent(self.TopLevel, stat_ev)
#                    time.sleep(5)

                    # Record room conditions
                    self.Troom = devices.ROLES_INSTR['GMHroom'].Measure('T')
                    self.Proom = devices.ROLES_INSTR['GMHroom'].Measure('P')
                    self.RHroom = devices.ROLES_INSTR['GMHroom'].Measure('RH')

                    self.WriteDataThisRow(row, node)
                    self.PlotThisRow(row, node)
                    time.sleep(0.1)
                    row += 1

                    # Reset start row for next measurement
#                    self.ws['B1'].value = row+3
                # (end of V3_mask loop)
            # (end of node loop)
        # (end of abs_V3 loop)

        self.FinishRun()
        return

    def SetNode(self, node):
        '''
        Update Node ComboBox and Change I/P node relays in IV-box
        '''
        print'AqnThread.SetNode(): ', node
        logger.info('Node = %s', node)
        self.RunPage.Node.SetValue(node)  # Update widget value
        s = node[1]
        if s in ('1', '2'):
            print'AqnThread.SetNode():Sending IVbox "', s, '"'
            logger.info('Sending IVbox "%s"', s)
            devices.ROLES_INSTR['IVbox'].SendCmd(s)
        else:  # '3'
            print'AqnThread.SetNode():IGNORING IVbox cmd "', s, '"'
            logger.info('IGNORING IVbox cmd "%s"', s)
        if not devices.ROLES_INSTR['IVbox'].demo:
            time.sleep(1)

    def PlotThisRow(self, row, node):
        # Plot data
        Dates = []
        for d in self.Times:
            Dates.append(dt.datetime.fromtimestamp(d))
        clear_plot = 0
#        if row == self.start_row:
        if row == 1:
            clear_plot = 1  # start each run with a clear plot

        plot_ev = evts.PlotEvent(t=Dates, V12=self.V12Data[node],
                                 V3=self.V3Data, clear=clear_plot, node=node)
        wx.PostEvent(self.PlotPage, plot_ev)

#    def WriteHeadings(self):
#        Id_row = self.start_row-2  # Headings
#        self.ws['A'+str(Id_row)].font = Font(b=True)
#        self.ws['A'+str(Id_row)] = 'Run ID:'
#        self.ws['B'+str(Id_row)].font = Font(b=True)
#        self.ws['B'+str(Id_row)] = self.RunPage.run_id
#
#        Head_row = self.start_row-1  # Headings
#
#        col_headings = {'A': 'Comment', 'B': 'DUC gain (V/A)', 'C': 'Rs (Ohm)',
#                        'D': 'I/P node (V1,V2?)', 'E': 'Date, time',
#                        'F': 'N meas.', 'G': 'Nom. Vout ', 'H': 'O/P V (V3)',
#                        'I': 'Stdev', 'J': 'I/P V (V1 or V2)', 'K': 'Stdev',
#                        'L': 'T(GMH)', 'M': 'Pt (DVM)', 'N': 'IP DVM range',
#                        'O': 'OP DVM range', 'P': 'T (room)', 'Q': 'P (room)',
#                        'R': 'RH (room)', 'S': 'Role',
#                        'T': 'Instrument description'}
#        for k in col_headings.keys():
#            self.ws[k+str(Head_row)] = col_headings[k]

    def WriteInstrAssignments(self):
        '''
        Record all roles and corresponding instrument descriptions in XL sheet
        '''
        role_row = self.start_row
        bord_tl = Border(top=Side(style='thin'), left=Side(style='thin'))
        bord_tr = Border(top=Side(style='thin'), right=Side(style='thin'))
        bord_l = Border(left=Side(style='thin'))
        bord_r = Border(right=Side(style='thin'))
        bord_bl = Border(bottom=Side(style='thin'), left=Side(style='thin'))
        bord_br = Border(bottom=Side(style='thin'), right=Side(style='thin'))
        for r in devices.ROLES_WIDGETS.keys():
            if role_row == self.start_row:  # 1st row
                self.ws['S'+str(role_row)].border = bord_tl
                self.ws['T'+str(role_row)].border = bord_tr
            elif role_row == self.start_row + 6:  # last row
                self.ws['S'+str(role_row)].border = bord_bl
                self.ws['T'+str(role_row)].border = bord_br
            else:  # in-between rows
                self.ws['S'+str(role_row)].border = bord_l
                self.ws['T'+str(role_row)].border = bord_r
            self.ws['S'+str(role_row)] = r
            d = devices.ROLES_WIDGETS[r]['icb'].GetValue()  # descr
            self.ws['T'+str(role_row)] = d
            role_row += 1

    def initialise(self):
        stat_ev = evts.StatusEvent(msg='Initialising instruments...', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)

        for r in devices.ROLES_INSTR.keys():
            if r == 'IVbox':
                d = 'IV_box'
            else:
                d = devices.ROLES_WIDGETS[r]['icb'].GetValue()
            # Open non-GMH devices:
            if 'GMH' not in devices.ROLES_INSTR[r].Descr:
                print'AqnThread.initialise(): Opening', d
                logger.info('Opening %s', d)
                devices.ROLES_INSTR[r].Open()
            else:
                print'AqnThread.initialise(): %s already open' % d
                logger.info('%s already open', d)

            stat_ev = evts.StatusEvent(msg=d, field=1)
            wx.PostEvent(self.TopLevel, stat_ev)
            devices.ROLES_INSTR[r].Init()
            if not devices.ROLES_INSTR[r].demo:
                time.sleep(1)
        stat_ev = evts.StatusEvent(msg='Done', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)

    def SetUpMeasThisRow(self, node):
        d = devices.ROLES_INSTR['SRC'].Descr
        if 'F5520A' in d:
            err = devices.ROLES_INSTR['SRC'].CheckErr()  # 'ERR?','*CLS'
            print'Cleared F5520A error:', err
            logger.info('Cleared F5520A error: %s', err)

        del self.V12Data[node][:]
        del self.V3Data[:]
        del self.Times[:]

    def MeasureV(self, node):
        assert node in ('V1', 'V2', 'V3'), 'Unknown argument to MeasureV().'

        if node == 'V1':
            if devices.ROLES_INSTR['DVM12'].demo is True:
                dvmOP = np.random.normal(self.V1_set,
                                         1.0e-5*abs(self.V1_set)+1e-6)
                self.V12Data['V1'].append(dvmOP)
            else:
                dvmOP = devices.ROLES_INSTR['DVM12'].Read()
                V = float(filter(self.filt, dvmOP))
                self.V12Data['V1'].append(V)

        elif node == 'V2':
            if devices.ROLES_INSTR['DVM12'].demo is True:
                dvmOP = np.random.normal(0.0,
                                         1.0e-5*abs(self.V1_set)+1e-6)
                self.V12Data['V2'].append(dvmOP)
            else:
                dvmOP = devices.ROLES_INSTR['DVM12'].Read()
                V = float(filter(self.filt, dvmOP))
                self.V12Data['V2'].append(V)

        elif node == 'V3':
            '''
            Just want one set of 20 timestamps -
            could have been either V1 or V2 instead.
            '''
            self.Times.append(time.time())
            if devices.ROLES_INSTR['DVM3'].demo is True:
                dvmOP = np.random.normal(self.Vout,
                                         1.0e-5*abs(self.Vout)+1e-6)
                self.V3Data.append(dvmOP)
            else:
                dvmOP = devices.ROLES_INSTR['DVM3'].Read()
                V = float(filter(self.filt, dvmOP))
                print'dvmOP =', V
                self.V3Data.append(V)
        if not (devices.ROLES_INSTR['DVM12'].demo and
                devices.ROLES_INSTR['DVM3'].demo):
                    time.sleep(0.1)
        return 1

    def WriteDataThisRow(self, row, node):
        stat_ev = evts.StatusEvent(msg='AqnThread.WriteDataThisRow():',
                                   field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        stat_ev = evts.StatusEvent(msg='Row '+str(row), field=1)
        wx.PostEvent(self.TopLevel, stat_ev)

#        self.ws['A'+str(row)] = self.Comment
#        self.ws['B'+str(row)] = self.DUC_G
#        self.ws['C'+str(row)] = self.Rs
#        self.ws['D'+str(row)] = node
#        self.ws['E'+str(row)] = str(dt.datetime.fromtimestamp(np.mean(self.Times)).strftime("%d/%m/%Y %H:%M:%S"))
#        self.ws['F'+str(row)] = NREADS
#        self.ws['G'+str(row)] = self.Vout  # Nominal output
#        self.ws['H'+str(row)] = self.V3m  # Measured output
#        self.ws['I'+str(row)] = self.V3sd
#        self.ws['J'+str(row)] = self.V12m[node]
#        self.ws['K'+str(row)] = self.V12sd[node]
#        self.ws['L'+str(row)] = self.T
        if devices.ROLES_INSTR['DVMT'].demo is True:
            TdvmOP = np.random.normal(108.0, 1.0e-2)
            self.TdvmOP = TdvmOP
        else:
            TdvmOP = devices.ROLES_INSTR['DVMT'].Read()  # .SendCmd('READ?')
            self.TdvmOP = float(filter(self.filt, TdvmOP))
#        self.ws['M'+str(row)] = self.TdvmOP
#        self.ws['N'+str(row)] = self.IPrange
#        self.ws['O'+str(row)] = self.OPrange
#        self.ws['P'+str(row)] = self.Troom
#        self.ws['Q'+str(row)] = self.Proom
#        self.ws['R'+str(row)] = self.RHroom
#
#        # Save after every row
#        self.wb_io.save(self.xlfilename)

        # Update run_dict:
        self.run_dict['Node'].append(node)
        self.run_dict['Nom_Vout'].append(self.Vout)
        self.run_dict['Date_time'].append(self.tm)
        self.run_dict['IP_V']['val'].append(self.V12m[node])
        self.run_dict['IP_V']['sd'].append(self.V12sd[node])
        self.run_dict['IPrange'].append(self.IPrange)
        self.run_dict['OP_V']['val'].append(self.V3m)
        self.run_dict['OP_V']['sd'].append(self.V3sd)
        self.run_dict['OPrange'].append(self.OPrange)
        self.run_dict['Pt_DVM'].append(self.TdvmOP)
        self.run_dict['T_GMH'].append(self.T)
        self.run_dict['Room_conds']['T'].append(self.Troom)
        self.run_dict['Room_conds']['P'].append(self.Proom)
        self.run_dict['Room_conds']['RH'].append(self.RHroom)

    def AbortRun(self):
        # prematurely end run, prompted by regular checks of _want_abort flag
        self.Standby()  # Set sources to 0V and leave system safe

        Update = {'progress': 100.0, 'end_flag': 1}
        stop_ev = evts.DataEvent(ud=Update)
        wx.PostEvent(self.RunPage, stop_ev)

        self.RunPage.StartBtn.Enable(True)
        self.RunPage.StopBtn.Enable(False)
        print'\nRun aborted.'
        logger.info('Run aborted.')

    def FinishRun(self):
        # Run complete - leave system safe and final xl save
#        self.wb_io.save(self.xlfilename)

        RunID = str(self.RunPage.run_id)
        data_file = str(self.RunPage.data_file)

        self.RunPage.master_run_dict.update({RunID: self.run_dict})
        with open(data_file, 'w') as IVY_out:
            logger.info('SAVING ALL RUN DATA...')
            json.dump(self.RunPage.master_run_dict, IVY_out)

        self.Standby()  # Set sources to 0V and leave system safe

        Update = {'progress': 100.0, 'end_flag': 1}
        stop_ev = evts.DataEvent(ud=Update)
        wx.PostEvent(self.RunPage, stop_ev)

        stat_ev = evts.StatusEvent(msg='RUN COMPLETED', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        stat_ev = evts.StatusEvent(msg='', field=1)
        wx.PostEvent(self.TopLevel, stat_ev)

        self.RunPage.StartBtn.Enable(True)
        self.RunPage.StopBtn.Enable(False)

    def Standby(self):
        # Set sources to 0V and disable outputs
        devices.ROLES_INSTR['SRC'].Stby  # .SendCmd('R0=')
        self.RunPage.V1Setting.SetValue(str(0))

    def abort(self):
        """abort worker thread."""
        # Method for use by main thread to signal an abort
        stat_ev = evts.StatusEvent(msg='abort(): Run aborted', field=0)
        wx.PostEvent(self.TopLevel, stat_ev)
        self._want_abort = 1
        time.sleep(1)

    def filt(self, char):
        '''
        A helper function to filter rubbish from DVM o/p unicode string
        and retain any number (as a string)
        '''
        accept_str = u'-0.12345678eE9'
        return char in accept_str  # Returns 'True' or 'False'


"""--------------End of Thread class definition-------------------"""
